from __future__ import annotations

import argparse
import csv
import re
from datetime import datetime
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

from .config import PipelineConfig, load_config
from .database import ResearchDatabase
from .sec_client import SecClient, sec_company_url, sec_companyfacts_api_url, sec_submission_api_url
from .workbook import WorkbookField, extract_company_seeds, inventory_workbooks


def run(config: PipelineConfig, sample_size: int | None = None, refresh_sec: bool = False) -> None:
    run_id = datetime.utcnow().strftime("%Y%m%d%H%M%S")
    workbook_paths = [
        path
        for key in ["master_workbook", "dashboard_workbook"]
        if config.inputs.get(key)
        for path in [config.path("inputs", key)]
        if path.exists()
    ]
    fields = inventory_workbooks(workbook_paths, config.path("outputs", "workbook_inventory"))
    seeds = extract_company_seeds(workbook_paths, config.path("inputs", "manual_company_seed"))

    client = SecClient(
        config.path("cache", "dir"),
        str(config.sec.get("user_agent") or "Langenberg research contact@example.com"),
        float(config.sec.get("request_sleep_seconds") or 0.12),
    )
    ticker_lookup = _ticker_lookup(client.load_company_tickers(str(config.cache["company_tickers_url"])))

    db = ResearchDatabase(config.path("outputs", "database"))
    db.start_run(run_id, f"seeded_from_workbooks={len(seeds)} sample_size={sample_size}")
    try:
        db.clear_derived_data()
        db.upsert_companies(seeds, ticker_lookup)
        companies = db.resolved_companies(sample_size)
        use_bulk = bool(config.sec.get("use_bulk")) and not sample_size
        submissions_zip = companyfacts_zip = None
        if use_bulk:
            submissions_zip = client.download(str(config.cache["submissions_bulk_url"]), "submissions.zip", refresh_sec)
            companyfacts_zip = client.download(str(config.cache["companyfacts_bulk_url"]), "companyfacts.zip", refresh_sec)
        forms = set(config.sec.get("forms") or ["10-K", "10-Q"])
        min_year = config.run.get("min_year")
        for company in companies:
            if use_bulk:
                submission = client.read_submission_from_zip(submissions_zip, company["cik"]) if submissions_zip else None
                facts = client.read_companyfacts_from_zip(companyfacts_zip, company["cik"]) if companyfacts_zip else None
            else:
                submission = client.get_json(sec_submission_api_url(company["cik"]))
                facts = client.get_json(sec_companyfacts_api_url(company["cik"]))
            db.insert_filings(_filing_rows(company, submission, forms, min_year))
            db.insert_raw_facts(_fact_rows(company, facts, forms, min_year))
        db.refresh_curated_metrics()
        validation_rows = _validation_rows(fields)
        db.write_validation(validation_rows)
        _write_validation_csv(validation_rows, config.path("outputs", "validation_export"))
        db.export_parquet(config.path("outputs", "parquet_dir"))
        export_excel(db, config.path("outputs", "excel_export"))
        db.finish_run(run_id, "success", f"companies={len(companies)}")
    except Exception as exc:
        db.finish_run(run_id, "failed", str(exc))
        raise
    finally:
        db.close()


def _ticker_lookup(rows: list[dict[str, Any]]) -> dict[str, dict[str, Any]]:
    lookup = {}
    for row in rows:
        ticker = str(row.get("ticker") or "").upper()
        if ticker:
            lookup[ticker] = row
            lookup[ticker.replace("-", ".")] = row
            lookup[ticker.replace(".", "-")] = row
    return lookup


def _filing_rows(company: dict, submission: dict | None, forms: set[str], min_year: int | None) -> list[list]:
    if not submission:
        return []
    recent = submission.get("filings", {}).get("recent", {})
    rows: list[list] = []
    for idx, form in enumerate(recent.get("form", [])):
        if form not in forms:
            continue
        filing_date = _date(_at(recent, "filingDate", idx))
        fiscal_year = _int(_at(recent, "fy", idx)) or (filing_date.year if filing_date else None)
        if min_year and fiscal_year and fiscal_year < min_year:
            continue
        accession = _at(recent, "accessionNumber", idx)
        rows.append([
            company["cik"],
            company["ticker"],
            company["name"],
            accession,
            form,
            _date(_at(recent, "reportDate", idx)),
            filing_date,
            fiscal_year,
            _at(recent, "fp", idx),
            _at(recent, "primaryDocument", idx),
            sec_company_url(company["cik"], accession),
        ])
    return rows


def _fact_rows(company: dict, facts: dict | None, forms: set[str], min_year: int | None) -> list[list]:
    if not facts:
        return []
    rows: list[list] = []
    facts_root = facts.get("facts", {})
    for taxonomy, concepts in facts_root.items():
        for concept, payload in concepts.items():
            label = payload.get("label")
            description = payload.get("description")
            for unit, values in payload.get("units", {}).items():
                for item in values:
                    form = item.get("form")
                    if form not in forms:
                        continue
                    fiscal_year = _int(item.get("fy"))
                    if min_year and fiscal_year and fiscal_year < min_year:
                        continue
                    accession = item.get("accn")
                    rows.append([
                        company["cik"],
                        company["ticker"],
                        company["name"],
                        taxonomy,
                        concept,
                        label,
                        description,
                        unit,
                        _float(item.get("val")),
                        accession,
                        form,
                        _date(item.get("filed")),
                        fiscal_year,
                        item.get("fp"),
                        _date(item.get("start")),
                        _date(item.get("end")),
                        item.get("frame"),
                        sec_company_url(company["cik"], accession),
                    ])
    return rows


def export_excel(db: ResearchDatabase, output_path: Path) -> None:
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    wb.remove(wb.active)
    queries = {
        "Company Universe": "SELECT * FROM companies ORDER BY ticker",
        "Latest Filing Snapshot": """
            SELECT * FROM filings
            QUALIFY row_number() OVER (PARTITION BY ticker, form ORDER BY filing_date DESC, accession_number DESC) = 1
            ORDER BY ticker, form
        """,
        "Metric History": """
            SELECT ticker, company_name, metric, fiscal_year, fiscal_period, period_end, value, unit, form, filed_date, source_url
            FROM curated_metrics
            ORDER BY ticker, metric, fiscal_year DESC, fiscal_period DESC
        """,
        "Period Comparison": """
            PIVOT (
                SELECT ticker, company_name, metric, fiscal_year, fiscal_period, value
                FROM curated_metrics
                WHERE metric IN ('revenue', 'operating_income', 'net_income', 'operating_cash_flow', 'free_cash_flow')
            )
            ON metric USING first(value)
            ORDER BY ticker, fiscal_year DESC, fiscal_period DESC
        """,
        "Raw Fact Lookup": """
            SELECT ticker, company_name, taxonomy, concept, unit, fiscal_year, fiscal_period, period_end, value, form, filed_date, source_url
            FROM raw_facts
            ORDER BY ticker, taxonomy, concept, fiscal_year DESC
            LIMIT 50000
        """,
        "Coverage Missing Data": """
            SELECT c.ticker, c.name, c.resolution_status, count(DISTINCT f.accession_number) AS filings,
                   count(rf.concept) AS raw_fact_rows,
                   count(DISTINCT cm.metric) AS curated_metrics
            FROM companies c
            LEFT JOIN filings f ON c.ticker = f.ticker
            LEFT JOIN raw_facts rf ON c.ticker = rf.ticker
            LEFT JOIN curated_metrics cm ON c.ticker = cm.ticker
            GROUP BY 1,2,3
            ORDER BY c.ticker
        """,
        "Source Links": """
            SELECT DISTINCT ticker, company_name, accession_number, form, filing_date, source_url
            FROM filings
            ORDER BY ticker, filing_date DESC
        """,
        "Workbook Validation": "SELECT * FROM workbook_validation ORDER BY workbook, sheet, row, column_index",
    }
    for sheet_name, query in queries.items():
        frame = db.con.execute(query).fetchdf()
        ws = wb.create_sheet(sheet_name[:31])
        ws.append(list(frame.columns))
        for row in frame.itertuples(index=False, name=None):
            ws.append(list(row))
        _format_sheet(ws)
    wb.save(output_path)


def _format_sheet(ws: Any) -> None:
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    for idx, column_cells in enumerate(ws.columns, start=1):
        values = [str(cell.value) for cell in column_cells[:200] if cell.value is not None]
        width = min(max([len(v) for v in values] + [10]) + 2, 45)
        ws.column_dimensions[get_column_letter(idx)].width = width


def _validation_rows(fields: list[WorkbookField]) -> list[dict]:
    status_map = {
        "enhanced_from_sec": "enhanced from SEC in V1",
        "mapped_from_workbook": "seed/mapping field retained",
        "future_non_sec_or_manual": "manual or future non-SEC source",
        "outside_sec_scope": "outside SEC company-data scope",
        "other": "inventory only",
    }
    return [
        {
            "workbook": f.workbook,
            "sheet": f.sheet,
            "row": f.row,
            "column_index": f.column,
            "value": f.value,
            "classification": f.classification,
            "v1_status": status_map.get(f.classification, "inventory only"),
        }
        for f in fields
    ]


def _write_validation_csv(rows: list[dict], path: Path) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=["workbook", "sheet", "row", "column_index", "value", "classification", "v1_status"])
        writer.writeheader()
        writer.writerows(rows)


def _at(data: dict[str, list], key: str, idx: int) -> Any:
    values = data.get(key) or []
    return values[idx] if idx < len(values) else None


def _date(value: Any) -> Any:
    if not value:
        return None
    return datetime.strptime(str(value)[:10], "%Y-%m-%d").date()


def _int(value: Any) -> int | None:
    try:
        return int(value)
    except (TypeError, ValueError):
        return None


def _float(value: Any) -> float | None:
    try:
        return float(value)
    except (TypeError, ValueError):
        return None


def main(argv: list[str] | None = None) -> None:
    parser = argparse.ArgumentParser(description="Build the U.S. SEC company research database.")
    parser.add_argument("--config", default="config/pipeline.yml")
    parser.add_argument("--sample-size", type=int, default=None)
    parser.add_argument("--refresh-sec", action="store_true")
    args = parser.parse_args(argv)
    config = load_config(args.config)
    run(config, args.sample_size, args.refresh_sec)
