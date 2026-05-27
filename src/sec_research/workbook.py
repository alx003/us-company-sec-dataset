from __future__ import annotations

import csv
import re
from dataclasses import dataclass
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


TICKER_RE = re.compile(r"^[A-Z][A-Z0-9.\-]{0,9}$")
NON_COMPANY_SYMBOLS = {
    "SPX", "DXY", "CNY", "EUR", "JPY", "GBP", "AUD", "CAD", "BRL", "INR",
    "BVSP", "GSPTSE", "MXX", "GDAXI", "RTSI", "SMI",
}


@dataclass(frozen=True)
class WorkbookField:
    workbook: str
    sheet: str
    row: int
    column: int
    value: str
    classification: str


@dataclass(frozen=True)
class CompanySeed:
    ticker: str
    cik: str | None
    name: str
    country: str | None
    sector: str | None
    industry: str | None
    source_workbook: str
    source_sheet: str
    source_row: int
    source_note: str


def inventory_workbooks(paths: list[Path], output_csv: Path) -> list[WorkbookField]:
    fields: list[WorkbookField] = []
    for path in paths:
        wb = load_workbook(path, read_only=True, data_only=True)
        for ws in wb.worksheets:
            for ridx, row in enumerate(ws.iter_rows(min_row=1, max_row=250, min_col=1, max_col=80), start=1):
                non_empty = [(cell.column, cell.value) for cell in row if _clean(cell.value)]
                if not non_empty:
                    continue
                text = " | ".join(_clean(value) for _, value in non_empty)
                classification = classify_field_text(text)
                if classification != "other" or len(non_empty) >= 3:
                    for column, value in non_empty[:30]:
                        fields.append(
                            WorkbookField(
                                workbook=path.name,
                                sheet=ws.title,
                                row=ridx,
                                column=column,
                                value=_clean(value),
                                classification=classify_field_text(_clean(value)),
                            )
                        )
    output_csv.parent.mkdir(parents=True, exist_ok=True)
    with output_csv.open("w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=WorkbookField.__dataclass_fields__.keys())
        writer.writeheader()
        for field in fields:
            writer.writerow(field.__dict__)
    return fields


def extract_company_seeds(workbook_paths: list[Path], manual_seed_csv: Path) -> list[CompanySeed]:
    seeds: dict[str, CompanySeed] = {}
    for path in workbook_paths:
        wb = load_workbook(path, read_only=True, data_only=True)
        for ws in wb.worksheets:
            header_row, header_map = _find_company_header(ws)
            if not header_row:
                continue
            max_row = min(ws.max_row, header_row + 800)
            for ridx, row in enumerate(
                ws.iter_rows(min_row=header_row + 1, max_row=max_row, min_col=1, max_col=80, values_only=True),
                start=header_row + 1,
            ):
                seed = _row_to_seed(path.name, ws.title, ridx, row, header_map)
                if seed and seed.ticker not in seeds:
                    seeds[seed.ticker] = seed

    if manual_seed_csv.exists():
        with manual_seed_csv.open(newline="", encoding="utf-8") as f:
            for row in csv.DictReader(f):
                ticker = _ticker(row.get("ticker"))
                name = _clean(row.get("name"))
                if not ticker or not name:
                    continue
                seeds[ticker] = CompanySeed(
                    ticker=ticker,
                    cik=_clean(row.get("cik")) or None,
                    name=name,
                    country=_clean(row.get("country")) or None,
                    sector=_clean(row.get("sector")) or None,
                    industry=_clean(row.get("industry")) or None,
                    source_workbook="manual_companies.csv",
                    source_sheet="manual_seed",
                    source_row=0,
                    source_note=_clean(row.get("source_note")),
                )
    return sorted(seeds.values(), key=lambda item: (item.country or "", item.ticker))


def classify_field_text(text: str) -> str:
    low = text.lower()
    if any(key in low for key in ["company data", "xbrl", "sec ", "edgar", "filing"]):
        return "enhanced_from_sec"
    if any(key in low for key in ["security", "symbol", "sector", "industry", "company", "name"]):
        return "mapped_from_workbook"
    if any(key in low for key in ["rss", "news", "linkedin", "email", "website", "youtube"]):
        return "future_non_sec_or_manual"
    if any(key in low for key in ["country", "gdp", "exports", "imports", "leader"]):
        return "outside_sec_scope"
    return "other"


def _find_company_header(ws: Any) -> tuple[int | None, dict[str, int]]:
    aliases = {
        "country": {"country", "nation"},
        "name": {"security", "company", "name"},
        "ticker": {"symbol", "ticker"},
        "sector": {"sector"},
        "industry": {"industry", "sub-industry", "sub industry"},
    }
    for ridx, row in enumerate(ws.iter_rows(min_row=1, max_row=250, min_col=1, max_col=80, values_only=True), start=1):
        values = [_clean(v).lower() for v in row]
        header_map: dict[str, int] = {}
        for field, candidates in aliases.items():
            for idx, value in enumerate(values):
                if value in candidates:
                    header_map[field] = idx
                    break
        if {"name", "ticker"}.issubset(header_map):
            return ridx, header_map
    return None, {}


def _row_to_seed(
    workbook: str,
    sheet: str,
    row_number: int,
    row: tuple[Any, ...],
    header_map: dict[str, int],
) -> CompanySeed | None:
    ticker = _ticker(_get(row, header_map.get("ticker")))
    name = _clean(_get(row, header_map.get("name")))
    if not name and _clean(_get(row, 2)):
        name = _clean(_get(row, 2))
    if not ticker or not name or ticker in NON_COMPANY_SYMBOLS:
        return None
    if len(name) < 3 or name.upper() == name and " " not in name:
        return None
    country = _clean(_get(row, header_map.get("country"))) or _clean(_get(row, 0)) or None
    if country == name:
        country = None
    sector = _clean(_get(row, header_map.get("sector"))) or None
    industry = _clean(_get(row, header_map.get("industry"))) or None
    if not sector or not industry:
        return None
    if name.lower().startswith(("s&p spdr", "spdr", "ftse", "hang seng")):
        return None
    return CompanySeed(
        ticker=ticker,
        cik=None,
        name=name,
        country=country,
        sector=sector,
        industry=industry,
        source_workbook=workbook,
        source_sheet=sheet,
        source_row=row_number,
        source_note="extracted from workbook company-like table",
    )


def _get(row: tuple[Any, ...], idx: int | None) -> Any:
    if idx is None or idx >= len(row):
        return None
    return row[idx]


def _ticker(value: Any) -> str:
    text = _clean(value).upper().replace("/", ".")
    return text if TICKER_RE.match(text) else ""


def _clean(value: Any) -> str:
    if value is None:
        return ""
    return " ".join(str(value).strip().split())
