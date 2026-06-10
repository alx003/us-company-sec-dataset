#!/usr/bin/env python3
from __future__ import annotations

import argparse
import csv
import gzip
import json
import os
import re
import time
import urllib.request
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

FORMS = {"10-K", "10-K/A", "10-Q", "10-Q/A"}
MIN_YEAR = 2018
DEFAULT_OUTPUT_DIR = Path.home() / "OneDrive" / "ImportantFiles" / "Brian" / "sec-workbooks"
USER_AGENT = os.environ.get("SEC_USER_AGENT", "SEC workbook automation contact@example.com")

METRIC_MAP = {
    "revenue": ["RevenueFromContractWithCustomerExcludingAssessedTax", "Revenues"],
    "gross_profit": ["GrossProfit"],
    "operating_income": ["OperatingIncomeLoss"],
    "net_income": ["NetIncomeLoss"],
    "eps": ["EarningsPerShareDiluted", "EarningsPerShareBasic"],
    "cash": ["CashAndCashEquivalentsAtCarryingValue"],
    "debt": ["DebtCurrent", "LongTermDebtCurrent", "LongTermDebtNoncurrent"],
    "operating_cash_flow": ["NetCashProvidedByUsedInOperatingActivities"],
    "capex": ["PaymentsToAcquirePropertyPlantAndEquipment"],
    "shares_outstanding": ["EntityCommonStockSharesOutstanding"],
    "rd": ["ResearchAndDevelopmentExpense"],
    "sga": ["SellingGeneralAndAdministrativeExpense"],
    "inventory": ["InventoryNet"],
    "assets": ["Assets"],
    "liabilities": ["Liabilities"],
    "equity": ["StockholdersEquity", "StockholdersEquityIncludingPortionAttributableToNoncontrollingInterest"],
}

FINANCIAL_STATEMENT_LAYOUT = [
    ("Income Statement", "revenue", "Revenue"),
    ("Income Statement", "gross_profit", "Gross Profit"),
    ("Income Statement", "operating_income", "Operating Income"),
    ("Income Statement", "net_income", "Net Income"),
    ("Income Statement", "eps", "Diluted EPS"),
    ("Cash Flow Statement", "operating_cash_flow", "Operating Cash Flow"),
    ("Cash Flow Statement", "capex", "Capex"),
    ("Cash Flow Statement", "free_cash_flow", "Free Cash Flow"),
    ("Balance Sheet", "cash", "Cash"),
    ("Balance Sheet", "debt", "Debt"),
    ("Balance Sheet", "inventory", "Inventory"),
    ("Balance Sheet", "assets", "Total Assets"),
    ("Balance Sheet", "liabilities", "Total Liabilities"),
    ("Balance Sheet", "equity", "Equity"),
    ("Other", "shares_outstanding", "Shares Outstanding"),
    ("Operating Expense", "rd", "R&D"),
    ("Operating Expense", "sga", "SG&A"),
]

SEGMENT_METRIC_PATTERNS = {
    "revenue": re.compile(r"revenue|sales", re.I),
    "profit": re.compile(r"profit|income|earnings|loss", re.I),
    "capex": re.compile(r"capex|capital|expenditure", re.I),
    "depreciation": re.compile(r"depreciation|amortization|depletion", re.I),
    "assets": re.compile(r"asset|propertyplant", re.I),
    "expense": re.compile(r"expense|cost", re.I),
}


def get_json(url: str) -> dict[str, Any]:
    request = urllib.request.Request(
        url,
        headers={"User-Agent": USER_AGENT, "Accept-Encoding": "gzip, deflate"},
    )
    with urllib.request.urlopen(request, timeout=90) as response:
        payload = response.read()
        if response.headers.get("Content-Encoding") == "gzip" or payload.startswith(b"\x1f\x8b"):
            payload = gzip.decompress(payload)
    time.sleep(0.15)
    return json.loads(payload.decode("utf-8"))


def get_text(url: str) -> str:
    request = urllib.request.Request(
        url,
        headers={"User-Agent": USER_AGENT, "Accept-Encoding": "gzip, deflate"},
    )
    with urllib.request.urlopen(request, timeout=90) as response:
        payload = response.read()
        if response.headers.get("Content-Encoding") == "gzip" or payload.startswith(b"\x1f\x8b"):
            payload = gzip.decompress(payload)
    time.sleep(0.15)
    return payload.decode("utf-8", errors="replace")


def sec_url(cik: str, accession: str | None = None) -> str:
    if not accession:
        return f"https://www.sec.gov/edgar/browse/?CIK={int(cik)}"
    return f"https://www.sec.gov/Archives/edgar/data/{int(cik)}/{accession.replace('-', '')}/"


def sec_document_url(cik: str, accession: str, primary_document: str) -> str:
    return f"{sec_url(cik, accession)}{primary_document}"


def load_ticker_lookup() -> dict[str, dict[str, Any]]:
    data = get_json("https://www.sec.gov/files/company_tickers_exchange.json")
    fields = data["fields"]
    lookup = {}
    for row in data["data"]:
        item = dict(zip(fields, row))
        ticker = str(item["ticker"]).upper()
        lookup[ticker] = item
        lookup[ticker.replace("-", ".")] = item
        lookup[ticker.replace(".", "-")] = item
    return lookup


def seed_tickers(path: Path | None) -> list[str]:
    if not path or not path.exists():
        return []
    tickers = []
    with path.open(newline="", encoding="utf-8") as f:
        for row in csv.DictReader(f):
            ticker = (row.get("ticker") or "").strip().upper()
            if ticker:
                tickers.append(ticker)
    return tickers


def at(data: dict[str, list[Any]], key: str, index: int) -> Any:
    values = data.get(key) or []
    return values[index] if index < len(values) else ""


def to_int(value: Any) -> int | None:
    try:
        return int(value)
    except (TypeError, ValueError):
        return None


def filing_rows(submissions: dict[str, Any], cik: str) -> list[dict[str, Any]]:
    recent = submissions.get("filings", {}).get("recent", {})
    rows = []
    for i, form in enumerate(recent.get("form", [])):
        fiscal_year = to_int(at(recent, "fy", i))
        if form not in FORMS or (fiscal_year and fiscal_year < MIN_YEAR):
            continue
        accession = at(recent, "accessionNumber", i)
        rows.append({
            "accession_number": accession,
            "form": form,
            "fiscal_year": fiscal_year,
            "fiscal_period": at(recent, "fp", i),
            "report_date": at(recent, "reportDate", i),
            "filing_date": at(recent, "filingDate", i),
            "primary_document": at(recent, "primaryDocument", i),
            "source_url": sec_url(cik, accession),
        })
    return rows


def raw_fact_rows(company: dict[str, Any], facts: dict[str, Any]) -> list[dict[str, Any]]:
    rows = []
    for taxonomy, concepts in facts.get("facts", {}).items():
        for concept, payload in concepts.items():
            for unit, values in payload.get("units", {}).items():
                for item in values:
                    fiscal_year = to_int(item.get("fy"))
                    if item.get("form") not in FORMS or (fiscal_year and fiscal_year < MIN_YEAR):
                        continue
                    accession = item.get("accn")
                    rows.append({
                        "ticker": company["ticker"],
                        "company_name": company["name"],
                        "taxonomy": taxonomy,
                        "concept": concept,
                        "label": payload.get("label", ""),
                        "unit": unit,
                        "value": item.get("val"),
                        "form": item.get("form"),
                        "fiscal_year": fiscal_year,
                        "fiscal_period": item.get("fp"),
                        "period_start": item.get("start", ""),
                        "period_end": item.get("end", ""),
                        "filed_date": item.get("filed", ""),
                        "accession_number": accession,
                        "source_url": sec_url(company["cik"], accession),
                    })
    return rows


def local_name(value: str) -> str:
    return str(value or "").split(":")[-1]


def clean_text(value: Any) -> str:
    text = re.sub(r"<[^>]+>", " ", str(value or ""))
    text = (
        text.replace("&nbsp;", " ")
        .replace("&amp;", "&")
        .replace("&lt;", "<")
        .replace("&gt;", ">")
    )
    return " ".join(text.split())


def tag_attrs(tag: str) -> dict[str, str]:
    attrs = {}
    for match in re.finditer(r'([\w:.-]+)\s*=\s*("([^"]*)"|\'([^\']*)\')', tag):
        attrs[match.group(1).lower()] = match.group(3) or match.group(4) or ""
    return attrs


def parse_number(text: str, scale: str = "0", sign: str = "") -> float | None:
    cleaned = clean_text(text).replace(",", "")
    if not cleaned or cleaned == "-":
        return None
    try:
        value = float(cleaned.replace("(", "").replace(")", ""))
    except ValueError:
        return None
    value *= 10 ** int(scale or 0)
    if sign == "-" or cleaned.startswith("("):
        value *= -1
    return value


def parse_contexts(html: str) -> dict[str, dict[str, Any]]:
    contexts = {}
    context_re = re.compile(r'<(?:xbrli:)?context\b[^>]*\bid\s*=\s*["\']([^"\']+)["\'][^>]*>(.*?)</(?:xbrli:)?context>', re.I | re.S)
    member_re = re.compile(r"<(?:xbrldi:)?explicitmember\b([^>]*)>(.*?)</(?:xbrldi:)?explicitmember>", re.I | re.S)
    for match in context_re.finditer(html):
        context_id = match.group(1)
        body = match.group(2)
        dimensions = []
        for member in member_re.finditer(body):
            attr = tag_attrs(member.group(1))
            dimensions.append(f"{local_name(attr.get('dimension', ''))}={local_name(clean_text(member.group(2)))}")
        start = re.search(r"<(?:xbrli:)?startdate>([^<]+)</(?:xbrli:)?startdate>", body, re.I)
        end = re.search(r"<(?:xbrli:)?enddate>([^<]+)</(?:xbrli:)?enddate>", body, re.I)
        instant = re.search(r"<(?:xbrli:)?instant>([^<]+)</(?:xbrli:)?instant>", body, re.I)
        contexts[context_id] = {
            "dimensions": dimensions,
            "period_start": start.group(1) if start else "",
            "period_end": (end.group(1) if end else "") or (instant.group(1) if instant else ""),
        }
    return contexts


def is_segment_dimension(dimensions: list[str]) -> bool:
    text = " ".join(dimensions).lower()
    if re.search(r"retirement|pension|benefitplan|definedbenefit|award|stockoption|sharebased", text):
        return False
    return bool(re.search(r"segment|geograph|region|product|service|business|domestic|international|express|freight|ground|office|aircraft|boeing", text))


def segment_metric(concept: str) -> str | None:
    for metric, pattern in SEGMENT_METRIC_PATTERNS.items():
        if pattern.search(concept):
            return metric
    return None


def business_segment(dimensions: str) -> str:
    low = dimensions.lower()
    if "federalexpress" in low or "federal express" in low:
        return "Federal Express"
    if "fedexfreight" in low or "freight" in low:
        return "FedEx Freight"
    if "fedexoffice" in low or "office" in low:
        return "FedEx Office"
    if "ground" in low:
        return "Ground"
    if "international" in low:
        return "International"
    if "domestic" in low:
        return "Domestic"
    return "Other / Tagged Detail"


def sector_bucket(dimensions: str, concept: str) -> str:
    text = f"{dimensions} {concept}".lower()
    if re.search(r"aircraft|boeing|md11|aviation|aerospace", text):
        return "Aerospace"
    if "freight" in text:
        return "Freight / LTL"
    if "express" in text:
        return "Express / Logistics"
    if "ground" in text:
        return "Ground / Parcel"
    if "office" in text:
        return "Office / Services"
    if "international" in text:
        return "International"
    if "domestic" in text:
        return "Domestic"
    return "Other"


def inline_segment_rows(company: dict[str, Any], filings: list[dict[str, Any]]) -> list[dict[str, Any]]:
    rows = []
    fact_re = re.compile(r"<ix:nonfraction\b([^>]*)>(.*?)</ix:nonfraction>", re.I | re.S)
    for filing in filings:
        primary_document = filing.get("primary_document") or ""
        accession = filing.get("accession_number") or ""
        if not primary_document or not accession:
            continue
        try:
            source = sec_document_url(company["cik"], accession, primary_document)
            html = get_text(source)
        except Exception as exc:
            rows.append({
                "business_segment": "Fetch Warning",
                "sector_bucket": "Fetch Warning",
                "metric": "download_error",
                "concept": "FilingDownloadError",
                "value": "",
                "unit": "",
                "form": filing.get("form", ""),
                "fiscal_year": filing.get("fiscal_year", "") or (context.get("period_end", "")[:4] if context.get("period_end") else ""),
                "fiscal_period": filing.get("fiscal_period", ""),
                "period_start": "",
                "period_end": filing.get("report_date", ""),
                "dimensions": str(exc),
                "source_url": source if "source" in locals() else "",
            })
            continue
        contexts = parse_contexts(html)
        for fact in fact_re.finditer(html):
            attr = tag_attrs(fact.group(1))
            context = contexts.get(attr.get("contextref", ""))
            if not context or not is_segment_dimension(context["dimensions"]):
                continue
            concept = local_name(attr.get("name", ""))
            metric = segment_metric(concept)
            if not metric:
                continue
            value = parse_number(fact.group(2), attr.get("scale", "0"), attr.get("sign", ""))
            if value is None:
                continue
            dimensions = "; ".join(context["dimensions"])
            rows.append({
                "business_segment": business_segment(dimensions),
                "sector_bucket": sector_bucket(dimensions, concept),
                "metric": metric,
                "concept": concept,
                "value": value,
                "unit": attr.get("unitref", ""),
                "form": filing.get("form", ""),
                "fiscal_year": filing.get("fiscal_year", ""),
                "fiscal_period": filing.get("fiscal_period", ""),
                "period_start": context.get("period_start", ""),
                "period_end": context.get("period_end", "") or filing.get("report_date", ""),
                "dimensions": dimensions,
                "source_url": source,
            })
    return rows


def curated_rows(raw_rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    concept_to_metric = {concept: metric for metric, concepts in METRIC_MAP.items() for concept in concepts}
    rows = []
    seen = set()
    for row in raw_rows:
        metric = concept_to_metric.get(row["concept"])
        if not metric:
            continue
        key = (row["ticker"], metric, row["fiscal_year"], row["fiscal_period"], row["period_end"], row["accession_number"])
        if key in seen:
            continue
        seen.add(key)
        rows.append({
            "ticker": row["ticker"],
            "company_name": row["company_name"],
            "metric": metric,
            "value": row["value"],
            "unit": row["unit"],
            "form": row["form"],
            "fiscal_year": row["fiscal_year"],
            "fiscal_period": row["fiscal_period"],
            "period_end": row["period_end"],
            "filed_date": row["filed_date"],
            "xbrl_concept": row["concept"],
            "source_url": row["source_url"],
        })
    return rows


def financial_statement_rows(clean_rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    period_order = {"FY": 0, "Q1": 1, "Q2": 2, "Q3": 3, "Q4": 4}
    periods = sorted(
        {
            (row["fiscal_year"], row["fiscal_period"])
            for row in clean_rows
            if row.get("fiscal_year") and row.get("fiscal_period") in period_order
        },
        key=lambda item: (item[0] or 0, period_order.get(item[1], 99)),
    )
    period_labels = [f"{fy} {fp}" for fy, fp in periods]
    value_lookup = {}
    for row in clean_rows:
        if row.get("fiscal_period") not in period_order:
            continue
        label = f"{row.get('fiscal_year')} {row.get('fiscal_period')}"
        key = (row.get("metric"), label)
        if key not in value_lookup or str(row.get("filed_date", "")) > str(value_lookup[key].get("filed_date", "")):
            value_lookup[key] = row

    rows = []
    for statement, metric, display_name in FINANCIAL_STATEMENT_LAYOUT:
        output = {"statement": statement, "metric": display_name, "source_metric": metric}
        for label in period_labels:
            output[label] = value_lookup.get((metric, label), {}).get("value", "")
        rows.append(output)
    return rows


def segment_summary_rows(segment_facts: list[dict[str, Any]]) -> list[dict[str, Any]]:
    grouped = {}
    for row in segment_facts:
        key = (
            row.get("business_segment", ""),
            row.get("sector_bucket", ""),
            row.get("metric", ""),
            row.get("fiscal_year", ""),
            row.get("fiscal_period", ""),
            row.get("period_end", ""),
            row.get("unit", ""),
        )
        grouped.setdefault(key, {"value": 0, "source_url": row.get("source_url", ""), "dimensions": row.get("dimensions", "")})
        try:
            grouped[key]["value"] += float(row.get("value") or 0)
        except (TypeError, ValueError):
            pass
    return [
        {
            "business_segment": key[0],
            "sector_bucket": key[1],
            "metric": key[2],
            "fiscal_year": key[3],
            "fiscal_period": key[4],
            "period_end": key[5],
            "unit": key[6],
            "value": values["value"],
            "dimensions_example": values["dimensions"],
            "source_url": values["source_url"],
        }
        for key, values in sorted(grouped.items(), key=lambda item: (str(item[0][3]), str(item[0][4]), item[0][0], item[0][1], item[0][2]))
    ]


def add_sheet(wb: Workbook, name: str, rows: list[dict[str, Any]]) -> None:
    ws = wb.create_sheet(name)
    headers = list(rows[0].keys()) if rows else ["message"]
    ws.append(headers)
    for row in rows:
        ws.append([row.get(header, "") for header in headers])
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    for idx, column in enumerate(ws.columns, start=1):
        values = [str(cell.value) for cell in column[:200] if cell.value is not None]
        width = min(max([len(value) for value in values] + [10]) + 2, 55)
        ws.column_dimensions[get_column_letter(idx)].width = width


def write_workbook(
    output_path: Path,
    company: dict[str, Any],
    filings: list[dict[str, Any]],
    raw_rows: list[dict[str, Any]],
    clean_rows: list[dict[str, Any]],
    segment_facts: list[dict[str, Any]],
) -> None:
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    wb.remove(wb.active)
    add_sheet(wb, "Company Info", [
        {"field": "ticker", "value": company["ticker"]},
        {"field": "company_name", "value": company["name"]},
        {"field": "cik", "value": company["cik"]},
        {"field": "sec_company_page", "value": sec_url(company["cik"])},
        {"field": "filing_rows", "value": len(filings)},
        {"field": "curated_metric_rows", "value": len(clean_rows)},
        {"field": "raw_fact_rows", "value": len(raw_rows)},
        {"field": "segment_fact_rows", "value": len(segment_facts)},
    ])
    add_sheet(wb, "Financial Statements", financial_statement_rows(clean_rows))
    add_sheet(wb, "Segment Summary", segment_summary_rows(segment_facts))
    add_sheet(wb, "Filings", filings)
    add_sheet(wb, "Curated Metrics", clean_rows)
    add_sheet(wb, "Segment Facts", segment_facts[:50000])
    add_sheet(wb, "Raw XBRL Facts", raw_rows[:50000])
    add_sheet(wb, "Research Notes", [
        {"date": "", "topic": "", "notes": "", "source": "", "follow_up": ""},
        {"date": "", "topic": "", "notes": "", "source": "", "follow_up": ""},
        {"date": "", "topic": "", "notes": "", "source": "", "follow_up": ""},
    ])
    wb.save(output_path)


def main() -> None:
    parser = argparse.ArgumentParser(description="Build Excel company workbooks from SEC/XBRL data.")
    parser.add_argument("--ticker", action="append", help="Ticker to build. Can be used multiple times.")
    parser.add_argument("--seed", default="data/seeds/manual_companies.csv", help="CSV with a ticker column.")
    parser.add_argument("--output-dir", default=str(DEFAULT_OUTPUT_DIR), help="Folder where Excel files are saved.")
    parser.add_argument("--limit", type=int, default=None, help="Optional max number of tickers to build.")
    args = parser.parse_args()

    tickers = [t.upper() for t in (args.ticker or [])]
    if not tickers:
        tickers = seed_tickers(Path(args.seed))
    if args.limit:
        tickers = tickers[:args.limit]
    if not tickers:
        raise SystemExit("No tickers found. Use --ticker FDX or provide data/seeds/manual_companies.csv.")

    lookup = load_ticker_lookup()
    output_dir = Path(args.output_dir).expanduser()
    for ticker in tickers:
        match = lookup.get(ticker)
        if not match:
            print(f"SKIP {ticker}: not found in SEC ticker file")
            continue
        company = {"ticker": ticker, "name": match["name"], "cik": str(match["cik"]).zfill(10)}
        submissions = get_json(f"https://data.sec.gov/submissions/CIK{company['cik']}.json")
        facts = get_json(f"https://data.sec.gov/api/xbrl/companyfacts/CIK{company['cik']}.json")
        filings = filing_rows(submissions, company["cik"])
        raw_rows = raw_fact_rows(company, facts)
        clean_rows = curated_rows(raw_rows)
        segment_facts = inline_segment_rows(company, filings)
        output_path = output_dir / f"{ticker}_sec_xbrl_workbook.xlsx"
        write_workbook(output_path, company, filings, raw_rows, clean_rows, segment_facts)
        print(f"WROTE {output_path}")


if __name__ == "__main__":
    main()
